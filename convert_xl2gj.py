import os
import openpyxl
import re
import json
import argparse
import logging
import csv
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

logging.basicConfig(filename="error.log", level=logging.ERROR, format="%(asctime)s - %(message)s")


def get_input_files(path: str) -> list[str]:
    """
    Получает список файлов (.xlsx и .csv) по заданному пути.
    Фильтруются временные файлы (имена, начинающиеся с '~$').
    """
    valid_ext = {".xlsx", ".csv"}
    if os.path.isfile(path):
        ext = os.path.splitext(path)[1].lower()
        if ext in valid_ext and not os.path.basename(path).startswith("~$"):
            return [path]
        else:
            return []
    elif os.path.isdir(path):
        return [
            os.path.join(path, f)
            for f in os.listdir(path)
            if os.path.splitext(f)[1].lower() in valid_ext and not f.startswith("~$")
        ]
    return []


def is_valid_coordinate(coord_str: str) -> bool:
    """
    Проверяет, соответствует ли строка заданному формату координат.
    Поддерживаются как точка, так и запятая в качестве десятичного разделителя.
    Примеры допустимых форматов:
      - "N55.5 E37.5"
      - "N55,5 E37,5"
      - "55.5 37.5"
      - "55,5 37,5"
    """
    pattern = r'^(?:N\d+[.,]\d+\sE\d+[.,]\d+|E\d+[.,]\d+\sN\d+[.,]\d+|\d+[.,]\d+\s\d+[.,]\d+)$'
    if not isinstance(coord_str, str):
        return False
    if not re.match(pattern, coord_str):
        return False
    try:
        parts = coord_str.split()
        # Замена запятых на точки для преобразования в float
        values = [
            float(part[1:].replace(',', '.')) if part[0] in "NE" else float(part.replace(',', '.'))
            for part in parts
        ]
        if len(values) != 2:
            return False
        # Если первая часть начинается с "E", то порядок координат: (долгота, широта),
        # иначе (широта, долгота) – для проверки поменяем местами при необходимости.
        lon, lat = values if parts[0][0] == "E" else reversed(values)
        return 0 <= lon <= 180 and 0 <= lat <= 90
    except ValueError:
        return False


def parse_coordinates(coord_str: str) -> tuple[float, float]:
    """
    Преобразует строку с координатами в кортеж (широта, долгота).
    Перед обработкой заменяет запятую на точку.
    Пример входа: "N55,5 E37,5" или "55,5 37,5".
    """
    if isinstance(coord_str, str):
        coord_str = coord_str.replace(',', '.')
    parts = coord_str.split()
    lat = None
    lon = None
    raw_values = []
    for part in parts:
        if part.startswith("N"):
            try:
                lat = float(part[1:].replace(',', '.'))
            except ValueError:
                raise ValueError(f"Неверное значение широты: {part}")
        elif part.startswith("E"):
            try:
                lon = float(part[1:].replace(',', '.'))
            except ValueError:
                raise ValueError(f"Неверное значение долготы: {part}")
        else:
            try:
                raw_values.append(float(part.replace(',', '.')))
            except ValueError:
                raise ValueError(f"Невозможно преобразовать '{part}' в число")
    if lat is None:
        if raw_values:
            lat = raw_values.pop(0)
        else:
            raise ValueError("Не указана широта")
    if lon is None:
        if raw_values:
            lon = raw_values.pop(0)
        else:
            raise ValueError("Не указана долгота")
    return (lat, lon)


def find_first_coordinate(sheet, keyword: str) -> tuple[int, int]:
    """
    Ищет первую ячейку с координатой, рядом с которой находится заголовок, начинающийся с заданного ключевого слова.
    Например, если рядом с координатой находится текст "Номер" или "Привязка".
    """
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        for col_idx, cell in enumerate(row, start=1):
            if cell and isinstance(cell, str) and is_valid_coordinate(cell):
                prev_cell = sheet.cell(row=max(1, row_idx - 1), column=1).value
                if isinstance(prev_cell, str) and prev_cell.startswith(keyword):
                    return row_idx, col_idx
    return -1, -1


def read_excel_coordinates(file_path: str, start_cell: str = None, cycle_check: bool = True) -> tuple[list[tuple[float, float]] | None, list[tuple[float, float]]]:
    """
    Извлекает координаты полигона и привязки из Excel-файла.
    Если указан start_cell, поиск начнётся с заданной ячейки, иначе – через find_first_coordinate.
    При включённом cycle_check проверяется инкрементность нумерации (получаемой из соседних ячеек).
    """
    polygon_coordinates = []
    anchor_coordinates = []
    numbering_list = []
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active

        if start_cell:
            col_letter, row_number = coordinate_from_string(start_cell)
            poly_row, poly_col = row_number, column_index_from_string(col_letter)
        else:
            poly_row, poly_col = find_first_coordinate(sheet, "Номер")
        
        if poly_row == -1:
            logging.error(f"Не найдены координаты полигона в {file_path}")
            return None, None

        current_row = poly_row
        while True:
            coord_cell = sheet.cell(row=current_row, column=poly_col).value
            if coord_cell and isinstance(coord_cell, str) and is_valid_coordinate(coord_cell):
                try:
                    coord = parse_coordinates(coord_cell)
                except Exception as e:
                    logging.error(f"Ошибка парсинга координат в {file_path} строка {current_row}: {e}")
                    return None, None
                polygon_coordinates.append(coord)
                if cycle_check:
                    # Чтение нумерации: значения из ячеек, находящихся на 4 и 3 позиции левее
                    num_cell1 = sheet.cell(row=current_row, column=poly_col - 4).value
                    num_cell2 = sheet.cell(row=current_row, column=poly_col - 3).value
                    try:
                        num1 = int(num_cell1) if num_cell1 is not None else None
                        num2 = int(num_cell2) if num_cell2 is not None else None
                    except ValueError:
                        logging.error(f"Невозможно преобразовать нумерацию в {file_path} строка {current_row}")
                        return None, None
                    if num1 is None or num2 is None:
                        logging.error(f"Отсутствует нумерация в {file_path} строка {current_row}")
                        return None, None
                    numbering_list.append((num1, num2))
                current_row += 1
            else:
                break

        if not polygon_coordinates:
            logging.error(f"В файле {file_path} не найдены координаты полигона.")
            return None, None

        if cycle_check:
            for i in range(len(numbering_list) - 1):
                if numbering_list[i][1] != numbering_list[i+1][0]:
                    logging.error(f"Нарушена последовательность нумерации в {file_path} на строке {poly_row + i}")
                    return None, None
            if numbering_list and (numbering_list[-1][1] != numbering_list[0][0]):
                logging.error(f"Нарушена циклическая нумерация в {file_path}: последний элемент {numbering_list[-1][1]} не соответствует первому {numbering_list[0][0]}")
                return None, None

        # Извлечение координат привязки
        anchor_row, anchor_col = find_first_coordinate(sheet, "Привязка")
        if anchor_row != -1:
            for row in sheet.iter_rows(min_row=anchor_row, min_col=anchor_col, values_only=True):
                if row and row[0] and isinstance(row[0], str) and is_valid_coordinate(str(row[0])):
                    try:
                        anchor_coordinates.append(parse_coordinates(str(row[0])))
                    except Exception as e:
                        logging.error(f"Ошибка парсинга координат привязки в {file_path}: {e}")
                        break
                else:
                    break

    except Exception as e:
        logging.error(f"Ошибка чтения {file_path}: {e}")
        return None, None

    return polygon_coordinates, anchor_coordinates


def read_csv_coordinates(file_path: str, csv_order: list[str]) -> tuple[list[tuple[float, float]] | None, list[tuple[float, float]]]:
    """
    Извлекает координаты полигона из CSV-файла.
    Ожидается, что в файле строки с разделителем ";".
    
    Порядок столбцов определяется параметром csv_order (список из 3 значений), где:
      - "n"  – номер точки,
      - "lat"– широта,
      - "lon"– долгота.
    
    Допустимые варианты порядка:
      - ["n", "lat", "lon"] – номер, широта, долгота (по умолчанию);
      - ["lat", "lon", "n"] – широта, долгота, номер;
      - ["lon", "lat", "n"] – долгота, широта, номер.
    
    Перед преобразованием заменяются запятые на точки.
    Также проверяется, что номера точек возрастают на 1 (за исключением возможного дублирования первой точки в конце).
    Если полигон зациклен (последняя точка совпадает с первой), то она удаляется.
    """
    polygon_coordinates = []
    numbering = []
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f, delimiter=";")
            rows = list(reader)
        if not rows:
            logging.error(f"CSV файл {file_path} пустой")
            return None, None

        for idx, row in enumerate(rows):
            if len(row) < 3:
                logging.error(f"CSV файл {file_path} содержит менее 3 столбцов: {row}")
                return None, None
            try:
                point_num = None
                lat = None
                lon = None
                for i, col_type in enumerate(csv_order):
                    value = row[i].strip()
                    if col_type == "n":
                        # Замена запятой на точку и преобразование в число (если вдруг номер записан с запятой)
                        point_num = int(float(value.replace(',', '.')))
                    elif col_type == "lat":
                        lat = float(value.replace(',', '.'))
                    elif col_type == "lon":
                        lon = float(value.replace(',', '.'))
                if point_num is None or lat is None or lon is None:
                    logging.error(f"Не удалось прочитать все необходимые значения в {file_path} строка {idx+1}")
                    return None, None
            except Exception as e:
                logging.error(f"Ошибка преобразования строки в CSV файле {file_path} строка {idx+1}: {row} -> {e}")
                return None, None

            numbering.append(point_num)
            polygon_coordinates.append((lat, lon))

        # Проверка последовательности нумерации
        if len(numbering) > 1:
            for i in range(1, len(numbering)):
                # Если последняя строка дублирует первую (закрытие полигона), допускаем это
                if i == len(numbering) - 1 and numbering[i] == numbering[0]:
                    continue
                if numbering[i] != numbering[i - 1] + 1:
                    logging.error(f"Нарушена последовательность нумерации в {file_path} на строке {i+1}: {numbering[i-1]} -> {numbering[i]}")
                    return None, None

        # Если полигон зациклен (последняя точка совпадает с первой), удаляем последнюю точку
        if len(polygon_coordinates) > 1:
            first = polygon_coordinates[0]
            last = polygon_coordinates[-1]
            if abs(first[0] - last[0]) < 1e-6 and abs(first[1] - last[1]) < 1e-6:
                polygon_coordinates.pop()

    except Exception as e:
        logging.error(f"Ошибка чтения CSV файла {file_path}: {e}")
        return None, None

    return polygon_coordinates, []  # Для CSV привязка не предусмотрена


def generate_geojson(file_path: str, polygon_coords: list[tuple[float, float]], anchor_coords: list[tuple[float, float]], output_dir: str, create_anchor: bool = True):
    """
    Генерирует GeoJSON-файл для полигона и, при необходимости, для привязки.
    Для полигона используется незацикленный список точек – первая точка добавляется в конец при генерации.
    """
    file_name = os.path.splitext(os.path.basename(file_path))[0]
    os.makedirs(output_dir, exist_ok=True)
    if polygon_coords:
        geojson = {
            "type": "FeatureCollection",
            "name": f"{file_name}{os.path.splitext(file_path)[1]}",
            "features": [
                {
                    "type": "Feature",
                    "properties": {"name": "0", "buffer": 0},
                    "geometry": {"type": "Polygon", "coordinates": [polygon_coords + [polygon_coords[0]]]}
                }
            ],
            "crs": {"type": "name", "properties": {"name": "urn:ogc:def:crs:OGC:1.3:CRS84"}}
        }
        output_path = os.path.join(output_dir, f"{file_name}.geojson")
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(geojson, f, ensure_ascii=False, indent=4)
        print(f"GeoJSON сохранен: {output_path}")
    if create_anchor and anchor_coords:
        anchor_output_path = os.path.join(output_dir, f"{file_name}_.geojson")
        anchor_geojson = {
            "type": "FeatureCollection",
            "name": f"{file_name}_.geojson",
            "features": [
                {
                    "type": "Feature",
                    "geometry": {"type": "MultiPoint", "coordinates": anchor_coords},
                    "properties": {"name": "0", "buffer": 0}
                }
            ],
            "crs": {"type": "name", "properties": {"name": "urn:ogc:def:crs:OGC:1.3:CRS84"}}
        }
        with open(anchor_output_path, "w", encoding="utf-8") as f:
            json.dump(anchor_geojson, f, ensure_ascii=False, indent=4)
        print(f"GeoJSON сохранен: {anchor_output_path}")


def main():
    parser = argparse.ArgumentParser(description="Конвертер Excel/CSV в GeoJSON")
    parser.add_argument("input", type=str, nargs="?", default=os.getcwd(),
                        help="Путь к каталогу с Excel/CSV файлами или к конкретному файлу")
    parser.add_argument("-o", "--output", type=str, default="result",
                        help="Каталог для сохранения GeoJSON")
    parser.add_argument("--no-cycle-check", action="store_true",
                        help="Отключить проверку цикличности нумерации (только для Excel)")
    parser.add_argument("--start-cell", type=str, default=None,
                        help="Адрес ячейки начала парсинга координат полигона в Excel (например, F19)")
    parser.add_argument("--enable-anchor-geojson", action="store_true",
                        help="Включить создание GeoJSON с координатами привязки (по умолчанию отключено)")
    parser.add_argument("--csv-column-order", type=str, default="n,lat,lon",
                        help=("Порядок столбцов в CSV через запятую. Допустимые варианты: "
                              "'n,lat,lon' (номер точки, широта, долгота), 'lat,lon,n' (широта, долгота, номер точки) или "
                              "'lon,lat,n' (долгота, широта, номер точки)"))
    args = parser.parse_args()

    # Обработка аргумента порядка столбцов для CSV
    csv_order = [token.strip().lower() for token in args.csv_column_order.split(",")]
    if len(csv_order) != 3 or set(csv_order) != {"n", "lat", "lon"}:
        print("Неверный формат --csv-column-order. Допустимые варианты: 'n,lat,lon', 'lat,lon,n' или 'lon,lat,n'")
        return

    files = get_input_files(args.input)
    if not files:
        print("Нет файлов .xlsx или .csv для обработки.")
        return

    for file in files:
        print(f"Обрабатываем: {file}")
        ext = os.path.splitext(file)[1].lower()
        if ext == ".xlsx":
            polygon_coords, anchor_coords = read_excel_coordinates(
                file,
                start_cell=args.start_cell,
                cycle_check=not args.no_cycle_check
            )
        elif ext == ".csv":
            polygon_coords, anchor_coords = read_csv_coordinates(file, csv_order)
        else:
            continue

        if polygon_coords is None:
            print(f"Ошибка в файле {file}. Пропускаем.")
            continue

        generate_geojson(
            file,
            polygon_coords,
            anchor_coords,
            args.output,
            create_anchor=args.enable_anchor_geojson
        )
    print("Обработка завершена!")


if __name__ == "__main__":
    main()
